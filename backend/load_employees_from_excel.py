"""
Load employees from the two Excel roster files into the database.

Files expected at project root (one level above this backend/ directory):
  - Current Employees list with emails.xlsx
  - Soft Allocated.xlsx

Run from the backend/ directory:
    python load_employees_from_excel.py
"""
from __future__ import annotations

import sys
from pathlib import Path

# Ensure the app package is importable
sys.path.insert(0, str(Path(__file__).parent))

import openpyxl
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.database.base import Base
from app.database.session import SessionLocal, engine
from app.models.employee import Employee

# Paths to both Excel files (at workspace root, one level above backend/)
ROOT = Path(__file__).resolve().parents[1]
CURRENT_EMP_XLSX = ROOT / "Current Employees list with emails.xlsx"
SOFT_ALLOC_XLSX = ROOT / "Soft Allocated.xlsx"


def _ensure_schema() -> None:
    """Create tables (including any new columns) if they don't exist yet."""
    Base.metadata.create_all(bind=engine)


def _split_name(full_name: str) -> tuple[str, str]:
    name = (full_name or "").strip()
    parts = name.split()
    if len(parts) <= 1:
        return name, ""
    return parts[0], " ".join(parts[1:])


def _coerce_str(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def load_current_employees(db: Session, seen_emails: set[str]) -> int:
    """
    Sheet columns (0-indexed):
      0: Employee Code
      1: Employee Full Name
      2: Years Of Experience
      3: Designation  (job_title)
      4: Group        (group_name)
      5: Email
    """
    wb = openpyxl.load_workbook(CURRENT_EMP_XLSX, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_code = _coerce_str(row[0])
        full_name = _coerce_str(row[1])
        years_exp = _coerce_str(row[2])
        designation = _coerce_str(row[3])
        group = _coerce_str(row[4])
        email = _coerce_str(row[5])

        if not emp_code or not email:
            continue
        email = email.lower()
        if email in seen_emails:
            continue  # de-duplicate

        first_name, last_name = _split_name(full_name or "")
        employee = Employee(
            employee_id=emp_code,
            first_name=first_name,
            last_name=last_name,
            email=email,
            department=designation,          # use Designation as department
            job_title=designation,
            group_name=group,
            years_experience=years_exp,
            work_location=None,              # not in this sheet
            is_active=True,
        )
        db.add(employee)
        seen_emails.add(email)
        count += 1

    db.commit()
    return count


def load_soft_allocated(db: Session, seen_emails: set[str]) -> int:
    """
    Sheet columns (0-indexed):
      0: EmpCode
      1: EmailId
      2: Employee Name
      3: CurrentDesignation  (job_title)
      4: Work Location       (work_location)
    """
    wb = openpyxl.load_workbook(SOFT_ALLOC_XLSX, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        emp_code = _coerce_str(row[0])
        email = _coerce_str(row[1])
        full_name = _coerce_str(row[2])
        designation = _coerce_str(row[3])
        location = _coerce_str(row[4])

        if not emp_code or not email:
            continue
        email = email.lower()
        if email in seen_emails:
            continue  # already loaded from first file

        first_name, last_name = _split_name(full_name or "")
        employee = Employee(
            employee_id=emp_code,
            first_name=first_name,
            last_name=last_name,
            email=email,
            department=designation,
            job_title=designation,
            group_name="Soft Allocated",
            years_experience=None,
            work_location=location,
            is_active=True,
        )
        db.add(employee)
        seen_emails.add(email)
        count += 1

    db.commit()
    return count


def seed_employees() -> None:
    print("Ensuring database schema is up to date…")
    _ensure_schema()

    db: Session = SessionLocal()
    try:
        print("Clearing existing employee records…")
        # Remove employees (cascade deletes related team_members / event_participants)
        db.execute(delete(Employee))
        db.commit()

        seen: set[str] = set()

        print(f"Loading from: {CURRENT_EMP_XLSX.name}")
        n1 = load_current_employees(db, seen)
        print(f"  → {n1} employees loaded")

        print(f"Loading from: {SOFT_ALLOC_XLSX.name}")
        n2 = load_soft_allocated(db, seen)
        print(f"  → {n2} employees loaded")

        total = db.query(Employee).count()
        print(f"\nDone. Total employees in database: {total}")
    finally:
        db.close()


if __name__ == "__main__":
    seed_employees()
