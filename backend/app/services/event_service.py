from __future__ import annotations

from datetime import datetime

from app.models.event import Event
from app.models.event_participant import EventParticipant
from app.models.event_team import EventTeam
from app.repositories.employee_repository import EmployeeRepository
from app.repositories.event_repository import EventRepository
from app.repositories.team_repository import TeamRepository
from app.schemas.common import PaginatedResponse
from app.schemas.event import EventCreate, EventRead
from app.utils.exceptions import APIException


class EventService:
    def __init__(self, repository: EventRepository, employee_repository: EmployeeRepository, team_repository: TeamRepository) -> None:
        self.repository = repository
        self.employee_repository = employee_repository
        self.team_repository = team_repository

    def list(self, page: int, page_size: int, search: str | None = None, status: str | None = None, start_after: datetime | None = None, end_before: datetime | None = None) -> PaginatedResponse[EventRead]:
        items = self.repository.list_filtered(search=search, status=status, start_after=start_after, end_before=end_before)
        total = len(items)
        start = (page - 1) * page_size
        end = start + page_size
        page_items = [self._to_read(item) for item in items[start:end]]
        pages = max((total + page_size - 1) // page_size, 1) if total else 1
        return PaginatedResponse(items=page_items, total=total, page=page, page_size=page_size, pages=pages)

    def get(self, event_id: int) -> EventRead:
        item = self.repository.get_by_id(event_id)
        if not item:
            raise APIException("Event not found", 404)
        return self._to_read(item)

    def create(self, data: EventCreate) -> EventRead:
        organizer = self.employee_repository.get_by_id(data.organizer_id)
        if not organizer:
            raise APIException("Organizer employee not found", 404)

        employee_ids = set(data.invited_employee_ids)
        team_ids = set(data.invited_team_ids)

        for team_id in team_ids:
            team = self.team_repository.get_by_id(team_id)
            if not team:
                raise APIException(f"Team {team_id} not found", 404)
            team_members = [member.employee_id for member in team.members]
            employee_ids.update(team_members)

        for employee_id in employee_ids:
            employee = self.employee_repository.get_by_id(employee_id)
            if not employee:
                raise APIException(f"Employee {employee_id} not found", 404)

        from datetime import timedelta
        import uuid

        proposed_times = []
        if data.is_recurring and data.recurrence_until:
            current_start = data.start_time
            current_end = data.end_time
            until_date = data.recurrence_until.date()
            allowed_days: set[int] = set(data.recurrence_days) if data.recurrence_days else set(range(7))

            while current_start.date() <= until_date:
                if current_start.weekday() in allowed_days:
                    proposed_times.append((current_start, current_end))
                current_start += timedelta(days=1)
                current_end += timedelta(days=1)

            if not proposed_times:
                raise APIException(
                    "No occurrences were created — the selected days of the week never fall within the chosen date range.",
                    400,
                )
        else:
            proposed_times.append((data.start_time, data.end_time))

        new_participants = set(employee_ids)
        new_participants.add(data.organizer_id)
        if data.owner_id:
            owner = self.employee_repository.get_by_id(data.owner_id)
            if not owner:
                raise APIException("Owner employee not found", 404)
            new_participants.add(data.owner_id)

        if not data.ignore_clashes:
            for p_start, p_end in proposed_times:
                overlapping = self.repository.db.query(Event).filter(
                    Event.start_time < p_end,
                    Event.end_time > p_start
                ).all()

                for overlap_event in overlapping:
                    overlap_participants = {p.employee_id for p in overlap_event.participants}
                    if overlap_event.organizer_id:
                        overlap_participants.add(overlap_event.organizer_id)
                    if overlap_event.owner_id:
                        overlap_participants.add(overlap_event.owner_id)
                    
                    clashing_people = new_participants.intersection(overlap_participants)
                    if clashing_people:
                        clash_id = next(iter(clashing_people))
                        emp = self.employee_repository.get_by_id(clash_id)
                        name = f"{emp.first_name} {emp.last_name}" if emp else f"ID {clash_id}"
                        raise APIException(f"Scheduling clash: {name} is already booked for '{overlap_event.title}' on {p_start.strftime('%b %d at %H:%M')}.", 409)

        if data.is_recurring and data.recurrence_until:
            current_start = data.start_time
            current_end = data.end_time
            until_date = data.recurrence_until.date()

            # If specific days were chosen, use that set; otherwise every day (0-6).
            allowed_days: set[int] = set(data.recurrence_days) if data.recurrence_days else set(range(7))

            first_saved_event = None
            series_id = str(uuid.uuid4())

            for p_start, p_end in proposed_times:
                event = Event(
                    title=data.title,
                    description=data.description,
                    start_time=p_start,
                    end_time=p_end,
                        location=data.location,
                        status=data.status,
                        training_catalog_id=data.training_catalog_id,
                        organizer_id=data.organizer_id,
                        meeting_link=data.meeting_link or None,
                        assignment_included=data.assignment_included,
                        series_id=series_id,
                    )
                saved = self.repository.add(event)
                if not first_saved_event:
                    first_saved_event = saved
                
                participant_rows = [
                    EventParticipant(event_id=saved.id, employee_id=emp_id)
                    for emp_id in sorted(employee_ids)
                ]
                self.repository.db.add_all(participant_rows)

                team_rows = [
                    EventTeam(event_id=saved.id, team_id=team_id)
                    for team_id in sorted(team_ids)
                ]
                self.repository.db.add_all(team_rows)
                self.repository.db.commit()

            return self._to_read(first_saved_event)
        else:
            event = Event(
                title=data.title,
                description=data.description,
                start_time=data.start_time,
                end_time=data.end_time,
                location=data.location,
                status=data.status,
                training_catalog_id=data.training_catalog_id,
                organizer_id=data.organizer_id,
                owner_id=data.owner_id,
                meeting_link=data.meeting_link or None,
                assignment_included=data.assignment_included,
            )
            saved = self.repository.add(event)

            participant_rows = [
                EventParticipant(event_id=saved.id, employee_id=employee_id)
                for employee_id in sorted(employee_ids)
            ]
            self.repository.db.add_all(participant_rows)

            team_rows = [
                EventTeam(event_id=saved.id, team_id=team_id)
                for team_id in sorted(team_ids)
            ]
            self.repository.db.add_all(team_rows)
            self.repository.db.commit()

            return self._to_read(saved)

    def update(self, event_id: int, data: EventCreate) -> EventRead:
        item = self.repository.get_by_id(event_id)
        if not item:
            raise APIException("Event not found", 404)

        update_data = data.model_dump(exclude={"invited_employee_ids", "invited_team_ids", "is_recurring", "recurrence_until", "recurrence_days", "ignore_clashes"})
        for field, value in update_data.items():
            setattr(item, field, value)

        employee_ids = set(data.invited_employee_ids)
        team_ids = set(data.invited_team_ids)
        for team_id in team_ids:
            team = self.team_repository.get_by_id(team_id)
            if not team:
                raise APIException(f"Team {team_id} not found", 404)
            team_members = [member.employee_id for member in team.members]
            employee_ids.update(team_members)

        for emp_id in employee_ids:
            if not self.employee_repository.get_by_id(emp_id):
                raise APIException(f"Employee {emp_id} not found", 404)

        new_participants = set(employee_ids)
        new_participants.add(data.organizer_id)
        if data.owner_id:
            owner = self.employee_repository.get_by_id(data.owner_id)
            if not owner:
                raise APIException("Owner employee not found", 404)
            new_participants.add(data.owner_id)

        if not data.ignore_clashes:
            overlapping = self.repository.db.query(Event).filter(
                Event.id != event_id,
                Event.start_time < data.end_time,
                Event.end_time > data.start_time
            ).all()

            for overlap_event in overlapping:
                overlap_participants = {p.employee_id for p in overlap_event.participants}
                if overlap_event.organizer_id:
                    overlap_participants.add(overlap_event.organizer_id)
                if getattr(overlap_event, "owner_id", None):
                    overlap_participants.add(overlap_event.owner_id)
                
                clashing_people = new_participants.intersection(overlap_participants)
                if clashing_people:
                    clash_id = next(iter(clashing_people))
                    emp = self.employee_repository.get_by_id(clash_id)
                    name = f"{emp.first_name} {emp.last_name}" if emp else f"ID {clash_id}"
                    raise APIException(f"Scheduling clash: {name} is already booked for '{overlap_event.title}' on {data.start_time.strftime('%b %d at %H:%M')}.", 409)

        existing_p = {p.employee_id: p for p in item.participants}
        for emp_id in employee_ids:
            if emp_id not in existing_p:
                self.repository.db.add(EventParticipant(event_id=event_id, employee_id=emp_id, status="invited"))
        for emp_id, p in existing_p.items():
            if emp_id not in employee_ids:
                self.repository.db.delete(p)

        existing_t = {t.team_id: t for t in item.invited_teams}
        for team_id in team_ids:
            if team_id not in existing_t:
                self.repository.db.add(EventTeam(event_id=event_id, team_id=team_id))
        for team_id, t in existing_t.items():
            if team_id not in team_ids:
                self.repository.db.delete(t)

        self.repository.db.commit()
        self.repository.db.refresh(item)
        return self._to_read(item)

    def delete(self, event_id: int, delete_series: bool = False) -> None:
        item = self.repository.get_by_id(event_id)
        if not item:
            raise APIException("Event not found", 404)
        
        if delete_series and item.series_id:
            # Delete all events in this series
            from app.models.event import Event
            events_in_series = self.repository.db.query(Event).filter(Event.series_id == item.series_id).all()
            for event in events_in_series:
                self.repository.delete(event)
        else:
            self.repository.delete(item)

    def upload_results(self, event_id: int, file_stream) -> None:
        import openpyxl
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        ws = wb.active

        # Read headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            if cell is not None:
                headers.append(str(cell).strip().lower())
            else:
                headers.append("")

        # Find column indices
        email_col = -1
        emp_id_col = -1
        attended_col = -1
        submitted_col = -1
        score_col = -1

        for idx, h in enumerate(headers):
            if "email" in h:
                email_col = idx
            elif "employee id" in h or "employee_id" in h or h == "id" or h == "employee code":
                emp_id_col = idx
            elif "attended" in h or "attendance" in h:
                attended_col = idx
            elif "submitted" in h or "submission" in h or "assignment" in h:
                submitted_col = idx
            elif "score" in h or "grade" in h or "points" in h:
                score_col = idx

        if email_col == -1 and emp_id_col == -1:
            raise APIException("Could not find email or employee ID column in Excel headers", 400)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and str(v).strip() for v in row):
                continue

            # Find employee
            employee = None
            if email_col != -1 and email_col < len(row) and row[email_col]:
                email = str(row[email_col]).strip()
                employee = self.employee_repository.get_by_email(email)
            if not employee and emp_id_col != -1 and emp_id_col < len(row) and row[emp_id_col]:
                emp_id = str(row[emp_id_col]).strip()
                from app.models.employee import Employee as EmpModel
                employee = self.repository.db.query(EmpModel).filter(EmpModel.employee_id == emp_id).first()

            if not employee:
                continue

            # Get or create participation
            p = self.repository.db.query(EventParticipant).filter(
                EventParticipant.event_id == event_id,
                EventParticipant.employee_id == employee.id
            ).first()

            if not p:
                p = EventParticipant(event_id=event_id, employee_id=employee.id, status="invited")
                self.repository.db.add(p)

            # Update fields
            if attended_col != -1 and attended_col < len(row) and row[attended_col] is not None:
                val = str(row[attended_col]).strip().lower()
                p.attended = val in {"yes", "y", "true", "1", "attended"}
                if p.attended:
                    p.status = "attended"

            if submitted_col != -1 and submitted_col < len(row) and row[submitted_col] is not None:
                val = str(row[submitted_col]).strip().lower()
                p.assignment_submitted = val in {"yes", "y", "true", "1", "submitted", "completed"}

            if score_col != -1 and score_col < len(row) and row[score_col] is not None:
                try:
                    p.assignment_score = int(float(str(row[score_col]).strip()))
                except (ValueError, TypeError):
                    p.assignment_score = None

        self.repository.db.commit()

    def _to_read(self, item: Event) -> EventRead:
        organizer_name = None
        if getattr(item, "organizer", None):
            organizer_name = f"{item.organizer.first_name} {item.organizer.last_name}"
        elif getattr(item, "organizer_id", None):
            organizer = self.employee_repository.get_by_id(item.organizer_id)
            if organizer:
                organizer_name = f"{organizer.first_name} {organizer.last_name}"

        owner_name = None
        if getattr(item, "owner", None):
            owner_name = f"{item.owner.first_name} {item.owner.last_name}"
        elif getattr(item, "owner_id", None):
            owner = self.employee_repository.get_by_id(item.owner_id)
            if owner:
                owner_name = f"{owner.first_name} {owner.last_name}"

        return EventRead(
            id=item.id,
            title=item.title,
            description=item.description,
            start_time=item.start_time,
            end_time=item.end_time,
            location=item.location,
            status=item.status,
            training_catalog_id=item.training_catalog_id,
            organizer_id=item.organizer_id,
            owner_id=item.owner_id,
            meeting_link=item.meeting_link,
            assignment_included=item.assignment_included,
            invited_employee_ids=[participant.employee_id for participant in item.participants],
            invited_team_ids=[invitation.team_id for invitation in item.invited_teams],
            organizer_name=organizer_name,
            owner_name=owner_name,
            invited_employee_names=[f"{p.employee.first_name} {p.employee.last_name}" for p in item.participants if p.employee],
            invited_team_names=[et.team.name for et in item.invited_teams if et.team],
            series_id=item.series_id,
        )
